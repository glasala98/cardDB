"""Generate a formatted Excel report from batch_price_output.csv.
Fair values are stored as real numbers so editing them on the All Cards
sheet automatically updates every summary, breakdown, and chart.
USD only — no CAD conversion.
"""
import csv, re
from collections import defaultdict
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.chart.series import SeriesLabel

# ── Config ─────────────────────────────────────────────────────────────────
DATA_START = 2      # first data row on All Cards
# DATA_END set after CSV load — exact last data row, no formula rows included

# Column indices on All Cards sheet
C_NUM    = 1   # A  row #
C_NAME   = 2   # B  card name
C_USD    = 3   # C  fair value USD  ← editable master
C_CONF   = 4   # D  confidence
C_SALES  = 5   # E  # sales
C_TOP3   = 6   # F  top 3 prices
C_MIN    = 7   # G  min
C_MAX    = 8   # H  max
C_NOTES  = 9   # I  notes           ← editable
C_PLAYER = 10  # J  player (helper for SUMIF)
C_SET    = 11  # K  set    (helper for SUMIF)

def col(idx): return get_column_letter(idx)
def ac_range(ci, s=None, e=None):
    """Absolute range on All Cards sheet for a given column index."""
    rs = DATA_START if s is None else s
    re_ = DATA_END   if e is None else e
    return f"'All Cards'!${col(ci)}${rs}:${col(ci)}${re_}"

# ── Styles ─────────────────────────────────────────────────────────────────
C_HEADER_BG  = "1F3864"; C_HEADER_FG = "FFFFFF"
C_HIGH       = "C6EFCE"; C_MEDIUM    = "DDEBF7"
C_LOW        = "FFEB9C"; C_ESTIMATED = "FCE4D6"
C_MANUAL     = "E2CFEA"; C_UNDER5    = "FFCCCC"
C_ALT        = "F2F2F2"; C_GOLD      = "FFF2CC"
C_HELPER     = "EFEFEF"; C_SUMMARY   = "2E75B6"
C_EDIT       = "FFFDE7"  # light yellow = editable cell hint

CONF_COLOR   = {"high": C_HIGH, "medium": C_MEDIUM, "low": C_LOW,
                "estimated": C_ESTIMATED, "manual": C_MANUAL}

def fill(h): return PatternFill("solid", fgColor=h)
thin = Side(style="thin", color="CCCCCC")
border = Border(left=thin, right=thin, top=thin, bottom=thin)
USD_FMT = '"$"#,##0.00'

def hdr(ws, row, text, bg, cols=3, fg=C_HEADER_FG, size=12):
    ws.merge_cells(f"A{row}:{col(cols)}{row}")
    c = ws.cell(row, 1, value=text)
    c.font = Font(name="Calibri", bold=True, size=size, color=fg)
    c.fill = fill(bg); c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[row].height = size + 10

# ── Helpers ─────────────────────────────────────────────────────────────────
def parse_price(val):
    if val is None: return None
    m = re.search(r"[\d.]+", str(val).replace(",", ""))
    return float(m.group()) if m else None

def player_from(card_name):
    parts = card_name.split(" - ")
    last = re.sub(r'\[.*?\]', '', parts[-1]).strip()
    last = re.sub(r'#\d+(?:/\d+)?', '', last).strip()
    last = re.sub(r'\(.*?\)', '', last).strip()
    m = re.match(r'^([A-Z][a-zA-Z\'\-]+(?:\s+[A-Z][a-zA-Z\'\-]+){1,2})', last)
    return m.group(1).strip() if m else last.strip()

def set_from(card_name):
    return card_name.split(" - ")[0].strip()

# ── Load CSV ────────────────────────────────────────────────────────────────
with open("batch_price_output.csv", encoding="utf-8", newline="") as f:
    rows = list(csv.DictReader(f))

# DATA_END = exact last data row — keeps SUMIF ranges away from any formula rows
DATA_END = len(rows) + DATA_START - 1

# Pre-compute for static sections that can't be fully dynamic
priced        = [r for r in rows if parse_price(r.get("Fair Value"))]
under5_rows   = sorted([r for r in priced if parse_price(r["Fair Value"]) < 5],
                       key=lambda r: parse_price(r["Fair Value"]))
manual_rows   = [r for r in rows if r.get("Confidence","").lower() == "manual"]

# For By Player / By Set ordering (formulas update values; Python sets initial order)
player_totals = defaultdict(float)
set_totals    = defaultdict(float)
for r in priced:
    p = player_from(r["Card Name"]); s = set_from(r["Card Name"])
    player_totals[p] += parse_price(r["Fair Value"])
    set_totals[s]    += parse_price(r["Fair Value"])

sorted_players = sorted(player_totals.items(), key=lambda x: -x[1])
sorted_sets    = sorted(set_totals.items(),    key=lambda x: -x[1])

tiers = [("$0 – $5",0,5),("$5 – $25",5,25),("$25 – $100",25,100),
         ("$100 – $500",100,500),("$500+",500,9_999_999)]

wb = Workbook()

# ════════════════════════════════════════════════════════════════════════════
# SHEET 1 — ALL CARDS  (master data — edit Fair Value here)
# ════════════════════════════════════════════════════════════════════════════
wc = wb.active
wc.title = "All Cards"
wc.sheet_view.showGridLines = False

col_headers = ["#","Card Name","Fair Value (USD) - EDIT HERE",
               "Confidence","# Sales","Top 3 Prices","Min","Max","Notes",
               "Player","Set"]
col_widths  = [5, 70, 22, 13, 7, 36, 10, 10, 28, 22, 45]

for ci, (h, w) in enumerate(zip(col_headers, col_widths), start=1):
    c = wc.cell(1, ci, value=h)
    c.font = Font(name="Calibri", bold=True, size=10, color=C_HEADER_FG)
    if ci == C_USD:
        c.fill = fill("C0A000")  # gold tint = editable
    elif ci in (C_PLAYER, C_SET):
        c.fill = fill("555555")  # dark = helper column
    else:
        c.fill = fill(C_HEADER_BG)
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=False)
    c.border = border
    wc.column_dimensions[col(ci)].width = w

wc.row_dimensions[1].height = 22
wc.freeze_panes = "C2"
wc.auto_filter.ref = f"A1:{col(C_SET)}1"

for i, row in enumerate(rows, start=DATA_START):
    price = parse_price(row.get("Fair Value"))
    conf  = row.get("Confidence","").lower()
    is_u5 = price is not None and price < 5
    row_bg  = C_UNDER5 if is_u5 else (C_ALT if i % 2 == 0 else "FFFFFF")
    conf_bg = CONF_COLOR.get(conf, "FFFFFF")
    player  = player_from(row["Card Name"])
    set_nm  = set_from(row["Card Name"])

    # Row number
    c = wc.cell(i, C_NUM, value=i - 1)
    c.fill = fill(row_bg); c.border = border
    c.font = Font(name="Calibri", size=10)
    c.alignment = Alignment(horizontal="center")

    # Card name
    c = wc.cell(i, C_NAME, value=row["Card Name"])
    c.fill = fill(row_bg); c.border = border
    c.font = Font(name="Calibri", size=10)

    # USD — stored as plain number, editable
    usd_cell = wc.cell(i, C_USD, value=price)
    usd_cell.number_format = USD_FMT
    usd_cell.fill = fill(C_EDIT)
    usd_cell.border = border
    usd_cell.font = Font(name="Calibri", size=10, bold=(price is not None and price >= 100))
    usd_cell.alignment = Alignment(horizontal="right")

    # Confidence
    conf_cell = wc.cell(i, C_CONF, value=row.get("Confidence",""))
    conf_cell.fill = fill(conf_bg); conf_cell.border = border
    conf_cell.font = Font(name="Calibri", size=10)
    conf_cell.alignment = Alignment(horizontal="center")

    # Sales, Top3, Min, Max
    for ci2, val in [(C_SALES, row.get("Num Sales","") or ""),
                     (C_TOP3,  row.get("Top 3 Prices","") or ""),
                     (C_MIN,   row.get("Min","") or ""),
                     (C_MAX,   row.get("Max","") or "")]:
        c = wc.cell(i, ci2, value=val)
        c.fill = fill(row_bg); c.border = border
        c.font = Font(name="Calibri", size=10)

    # Notes — editable
    c = wc.cell(i, C_NOTES, value="")
    c.fill = fill(C_EDIT); c.border = border
    c.font = Font(name="Calibri", size=10)

    # Helper columns — Player, Set
    for ci2, val in [(C_PLAYER, player), (C_SET, set_nm)]:
        c = wc.cell(i, ci2, value=val)
        c.fill = fill(C_HELPER); c.border = border
        c.font = Font(name="Calibri", size=9, color="666666")

# No grand total row on All Cards — avoids any formula inside the SUMIF range.
# Grand total lives on the Summary sheet.

# ════════════════════════════════════════════════════════════════════════════
# SHEET 2 — SUMMARY  (all formulas referencing All Cards)
# ════════════════════════════════════════════════════════════════════════════
ws = wb.create_sheet("Summary")
ws.sheet_view.showGridLines = False
ws.column_dimensions["A"].width = 36
ws.column_dimensions["B"].width = 22
ws.column_dimensions["C"].width = 40

hdr(ws, 1, "CARD COLLECTION — PRICE REPORT", C_HEADER_BG, cols=3, size=20)
ws.row_dimensions[1].height = 48
ws.row_dimensions[2].height = 8

def sum_row(ws, row, label, formula, bold=False, bg=None):
    a = ws.cell(row, 1, value=label)
    b = ws.cell(row, 2, value=formula)
    a.font = Font(name="Calibri", bold=bold, size=11)
    b.font = Font(name="Calibri", bold=bold, size=11)
    b.number_format = USD_FMT
    b.alignment = Alignment(horizontal="right")
    if bg:
        a.fill = fill(bg); b.fill = fill(bg)
    a.border = border; b.border = border
    ws.row_dimensions[row].height = 18

def count_row(ws, row, label, formula, bold=False, bg=None):
    a = ws.cell(row, 1, value=label)
    b = ws.cell(row, 2, value=formula)
    a.font = Font(name="Calibri", bold=bold, size=11)
    b.font = Font(name="Calibri", bold=bold, size=11)
    b.alignment = Alignment(horizontal="right")
    if bg:
        a.fill = fill(bg); b.fill = fill(bg)
    a.border = border; b.border = border
    ws.row_dimensions[row].height = 18

r = 3
hdr(ws, r, "OVERVIEW", C_SUMMARY, cols=3, size=12); r += 1
count_row(ws, r, "Total Cards in List",  len(rows)); r += 1
count_row(ws, r, "Cards Priced",
          f"=COUNTA({ac_range(C_USD)})-COUNTIF({ac_range(C_USD)},0)"); r += 1
count_row(ws, r, "Manual Entries",
          f'=COUNTIF({ac_range(C_CONF)},"manual")', bg=C_MANUAL); r += 1
count_row(ws, r, "Cards Under $5",
          f'=COUNTIFS({ac_range(C_USD)},">=0.01",{ac_range(C_USD)},"<5")', bg=C_UNDER5); r += 1

r += 1
hdr(ws, r, "VALUATION  (edit Fair Value on 'All Cards' sheet to update)",
    C_SUMMARY, cols=3, size=11); r += 1
sum_row(ws, r, "GRAND TOTAL",
        f"=SUM({ac_range(C_USD)})", bold=True, bg=C_GOLD); r += 1
sum_row(ws, r, "  Scraped (eBay)",
        f'=SUMIF({ac_range(C_CONF)},"<>manual",{ac_range(C_USD)})'); r += 1
sum_row(ws, r, "  Manual Comps",
        f'=SUMIF({ac_range(C_CONF)},"manual",{ac_range(C_USD)})', bg=C_MANUAL); r += 1

r += 1
hdr(ws, r, "VALUE BY CONFIDENCE", C_SUMMARY, cols=3, size=12); r += 1
for tier, bg in [("high",C_HIGH),("medium",C_MEDIUM),("low",C_LOW),
                 ("estimated",C_ESTIMATED),("manual",C_MANUAL)]:
    n = sum(1 for rw in rows if rw.get("Confidence","").lower() == tier)
    sum_row(ws, r, f"{tier.title()} ({n} cards)",
            f'=SUMIF({ac_range(C_CONF)},"{tier}",{ac_range(C_USD)})', bg=bg); r += 1

r += 1
hdr(ws, r, "VALUE TIERS", C_SUMMARY, cols=3, size=12); r += 1
ws.cell(r, 2, value="# Cards").font = Font(name="Calibri", bold=True, size=10, color=C_HEADER_FG)
ws.cell(r, 3, value="Total USD").font = Font(name="Calibri", bold=True, size=10, color=C_HEADER_FG)
for ci in [1,2,3]:
    ws.cell(r, ci).fill = fill(C_HEADER_BG)
    ws.cell(r, ci).border = border
ws.cell(r, 3).alignment = Alignment(horizontal="right")
ws.row_dimensions[r].height = 15; r += 1
for label, lo, hi in tiers:
    hi_s = str(hi) if hi < 9_000_000 else "9999999"
    count_f = f'=COUNTIFS({ac_range(C_USD)},">={lo}",{ac_range(C_USD)},"<{hi_s}")'
    sum_f   = f'=SUMIFS({ac_range(C_USD)},{ac_range(C_USD)},">={lo}",{ac_range(C_USD)},"<{hi_s}")'
    a = ws.cell(r, 1, value=label)
    b = ws.cell(r, 2, value=count_f)
    c = ws.cell(r, 3, value=sum_f)
    bg = C_ALT if r % 2 == 0 else "FFFFFF"
    for cell in [a, b, c]:
        cell.fill = fill(bg); cell.border = border
        cell.font = Font(name="Calibri", size=10)
    b.alignment = Alignment(horizontal="center")
    c.alignment = Alignment(horizontal="right")
    c.number_format = USD_FMT
    ws.row_dimensions[r].height = 16; r += 1

# Top 10 — values link to All Cards so they update when price is edited
r += 1
hdr(ws, r, "TOP 10 MOST VALUABLE  (snapshot ranking — values update live)",
    "C00000", cols=3, size=12); r += 1
for ci, h in enumerate(["Card Name","USD","Confidence"], start=1):
    c = ws.cell(r, ci, value=h)
    c.font = Font(name="Calibri", bold=True, size=10, color=C_HEADER_FG)
    c.fill = fill(C_HEADER_BG); c.border = border
ws.row_dimensions[r].height = 16; r += 1
sorted_priced = sorted(priced, key=lambda rw: parse_price(rw["Fair Value"]), reverse=True)
for rw in sorted_priced[:10]:
    ac_row = next((i + DATA_START for i, row in enumerate(rows)
                   if row["Card Name"] == rw["Card Name"]), None)
    conf = rw.get("Confidence","")
    bg   = CONF_COLOR.get(conf, "FFFFFF")
    ws.cell(r, 1, value=rw["Card Name"]).fill = fill(bg)
    if ac_row:
        usd_c = ws.cell(r, 2, value=f"='All Cards'!{col(C_USD)}{ac_row}")
        usd_c.number_format = USD_FMT
        usd_c.fill = fill(bg)
        usd_c.alignment = Alignment(horizontal="right")
    ws.cell(r, 3, value=conf.title()).fill = fill(bg)
    ws.cell(r, 3).alignment = Alignment(horizontal="center")
    for ci in [1, 2, 3]:
        ws.cell(r, ci).font = Font(name="Calibri", size=10, bold=(ci==2))
        ws.cell(r, ci).border = border
    ws.row_dimensions[r].height = 16; r += 1

# Legend
r += 1
hdr(ws, r, "CONFIDENCE LEGEND", C_HEADER_BG, cols=3); r += 1
legend = [("High",    C_HIGH,      "Direct eBay comp — full query matched variant + serial + set"),
          ("Medium",  C_MEDIUM,    "Set-level comp — dropped parallel/variant name"),
          ("Low",     C_LOW,       "Broad fallback — player + card# + year only"),
          ("Estimated",C_ESTIMATED,"Nearby print-run comp adjusted by serial multiplier"),
          ("Manual",  C_MANUAL,    "No eBay sales — price entered manually"),
          ("Under $5",C_UNDER5,    "Low-value flag — red rows on All Cards sheet")]
for tier, bg, desc in legend:
    ws.cell(r, 1, value=tier).fill = fill(bg)
    ws.merge_cells(f"B{r}:C{r}")
    ws.cell(r, 2, value=desc).fill = fill(bg)
    for ci in [1, 2]:
        ws.cell(r, ci).font = Font(name="Calibri", bold=(ci==1), size=10)
        ws.cell(r, ci).border = border
    ws.row_dimensions[r].height = 15; r += 1

# ════════════════════════════════════════════════════════════════════════════
# SHEET 3 — BY PLAYER  (SUMIF formulas — live)
# ════════════════════════════════════════════════════════════════════════════
wp = wb.create_sheet("By Player")
wp.sheet_view.showGridLines = False
wp.column_dimensions["A"].width = 30
wp.column_dimensions["B"].width = 10
wp.column_dimensions["C"].width = 18

hdr(wp, 1, "VALUE BY PLAYER  (totals update when you edit All Cards)", C_HEADER_BG, cols=3, size=13)
wp.row_dimensions[1].height = 30
wp.merge_cells("A2:C2"); wp.row_dimensions[2].height = 14
note = wp.cell(2, 1, value="Edit Fair Value on 'All Cards' sheet (yellow column) — totals here update automatically")
note.font = Font(name="Calibri", italic=True, size=9, color="666666")

for ci, h in enumerate(["Player","# Cards","Total USD"], start=1):
    c = wp.cell(3, ci, value=h)
    c.font = Font(name="Calibri", bold=True, size=11, color=C_HEADER_FG)
    c.fill = fill(C_SUMMARY); c.alignment = Alignment(horizontal="center"); c.border = border
wp.row_dimensions[3].height = 18
wp.freeze_panes = "A4"; wp.auto_filter.ref = "A3:C3"

for i, (player, _) in enumerate(sorted_players, start=4):
    bg = C_ALT if i % 2 == 0 else "FFFFFF"
    wp.cell(i, 1, value=player)
    cnt = wp.cell(i, 2, value=f'=COUNTIF({ac_range(C_PLAYER)},A{i})')
    usd = wp.cell(i, 3, value=f'=SUMIF({ac_range(C_PLAYER)},A{i},{ac_range(C_USD)})')
    usd.number_format = USD_FMT
    for ci in [1, 2, 3]:
        c = wp.cell(i, ci)
        c.fill = fill(bg); c.border = border
        c.font = Font(name="Calibri", size=10)
    cnt.alignment = Alignment(horizontal="center")
    usd.alignment = Alignment(horizontal="right")

pr = len(sorted_players) + 4
c = wp.cell(pr, 1, value="GRAND TOTAL")
c.font = Font(name="Calibri", bold=True, size=11)
tot_usd = wp.cell(pr, 3, value=f"=SUM({ac_range(C_USD)})")
tot_usd.number_format = USD_FMT
for ci in [1, 3]:
    c = wp.cell(pr, ci); c.fill = fill(C_GOLD); c.border = border
    c.font = Font(name="Calibri", bold=True, size=11)
tot_usd.alignment = Alignment(horizontal="right")

# ════════════════════════════════════════════════════════════════════════════
# SHEET 4 — BY SET  (SUMIF formulas — live)
# ════════════════════════════════════════════════════════════════════════════
wset = wb.create_sheet("By Set")
wset.sheet_view.showGridLines = False
wset.column_dimensions["A"].width = 52
wset.column_dimensions["B"].width = 10
wset.column_dimensions["C"].width = 18

hdr(wset, 1, "VALUE BY SET  (totals update when you edit All Cards)", C_HEADER_BG, cols=3, size=13)
wset.row_dimensions[1].height = 30
wset.merge_cells("A2:C2"); wset.row_dimensions[2].height = 14
note2 = wset.cell(2, 1, value="Edit Fair Value on 'All Cards' sheet (yellow column) — totals here update automatically")
note2.font = Font(name="Calibri", italic=True, size=9, color="666666")

for ci, h in enumerate(["Set Name","# Cards","Total USD"], start=1):
    c = wset.cell(3, ci, value=h)
    c.font = Font(name="Calibri", bold=True, size=11, color=C_HEADER_FG)
    c.fill = fill(C_SUMMARY); c.alignment = Alignment(horizontal="center"); c.border = border
wset.row_dimensions[3].height = 18
wset.freeze_panes = "A4"; wset.auto_filter.ref = "A3:C3"

for i, (set_name, _) in enumerate(sorted_sets, start=4):
    bg = C_ALT if i % 2 == 0 else "FFFFFF"
    wset.cell(i, 1, value=set_name)
    cnt = wset.cell(i, 2, value=f'=COUNTIF({ac_range(C_SET)},A{i})')
    usd = wset.cell(i, 3, value=f'=SUMIF({ac_range(C_SET)},A{i},{ac_range(C_USD)})')
    usd.number_format = USD_FMT
    for ci in [1, 2, 3]:
        c = wset.cell(i, ci)
        c.fill = fill(bg); c.border = border; c.font = Font(name="Calibri", size=10)
    cnt.alignment = Alignment(horizontal="center")
    usd.alignment = Alignment(horizontal="right")

sr = len(sorted_sets) + 4
c = wset.cell(sr, 1, value="GRAND TOTAL")
c.font = Font(name="Calibri", bold=True, size=11)
st_usd = wset.cell(sr, 3, value=f"=SUM({ac_range(C_USD)})")
st_usd.number_format = USD_FMT
for ci in [1, 3]:
    c = wset.cell(sr, ci); c.fill = fill(C_GOLD); c.border = border
    c.font = Font(name="Calibri", bold=True, size=11)
st_usd.alignment = Alignment(horizontal="right")

# ════════════════════════════════════════════════════════════════════════════
# SHEET 5 — CHARTS  (data tables use formulas; charts auto-update)
# ════════════════════════════════════════════════════════════════════════════
wch = wb.create_sheet("Charts")
wch.sheet_view.showGridLines = False
hdr(wch, 1, "VISUAL ANALYSIS  (charts update when Fair Value is edited)", C_HEADER_BG, cols=16, size=13)

# ── Confidence pie data table (SUMIF formulas) ─────────────────────────────
wch.cell(3, 1, "Confidence").font = Font(bold=True)
wch.cell(3, 2, "Value USD").font  = Font(bold=True)
tier_order = ["high","medium","low","estimated","manual"]
for k, t in enumerate(tier_order, start=4):
    wch.cell(k, 1, value=t.title())
    wch.cell(k, 2, value=f'=SUMIF({ac_range(C_CONF)},"{t}",{ac_range(C_USD)})')
    wch.cell(k, 2).number_format = USD_FMT

pie = PieChart(); pie.title = "Value by Confidence Tier (USD)"
pie.style = 10; pie.width = 14; pie.height = 12
pie_data = Reference(wch, min_col=2, min_row=3, max_row=8)
pie_cats = Reference(wch, min_col=1, min_row=4, max_row=8)
pie.add_data(pie_data, titles_from_data=True); pie.set_categories(pie_cats)
wch.add_chart(pie, "D3")

# ── Value tier bar (COUNTIFS formulas) ────────────────────────────────────
wch.cell(3, 15, "Tier").font   = Font(bold=True)
wch.cell(3, 16, "# Cards").font = Font(bold=True)
for k, (label, lo, hi) in enumerate(tiers, start=4):
    hi_s = str(hi) if hi < 9_000_000 else "9999999"
    wch.cell(k, 15, value=label)
    wch.cell(k, 16, value=f'=COUNTIFS({ac_range(C_USD)},">={lo}",{ac_range(C_USD)},"<{hi_s}")')

bar3 = BarChart(); bar3.title = "Cards by Value Tier"
bar3.y_axis.title = "# Cards"; bar3.style = 10; bar3.width = 14; bar3.height = 12
d3 = Reference(wch, min_col=16, min_row=3, max_row=8)
c3 = Reference(wch, min_col=15, min_row=4, max_row=8)
bar3.add_data(d3, titles_from_data=True); bar3.set_categories(c3)
bar3.series[0].graphicalProperties.pattFill = None
wch.add_chart(bar3, "P3")

# ── Top 10 bar (reference All Cards USD cells) ────────────────────────────
wch.cell(20, 1, "Player / Card").font = Font(bold=True)
wch.cell(20, 2, "USD").font = Font(bold=True)
for j, rw in enumerate(sorted_priced[:10], start=21):
    ac_row = next((i + DATA_START for i, row in enumerate(rows)
                   if row["Card Name"] == rw["Card Name"]), None)
    wch.cell(j, 1, value=player_from(rw["Card Name"]))
    if ac_row:
        wch.cell(j, 2, value=f"='All Cards'!{col(C_USD)}{ac_row}")
        wch.cell(j, 2).number_format = USD_FMT

bar = BarChart(); bar.type = "bar"
bar.title = "Top 10 Most Valuable Cards (USD)"
bar.y_axis.title = "Card"; bar.x_axis.title = "USD"
bar.style = 10; bar.width = 18; bar.height = 12
d1 = Reference(wch, min_col=2, min_row=20, max_row=30)
c1 = Reference(wch, min_col=1, min_row=21, max_row=30)
bar.add_data(d1, titles_from_data=True); bar.set_categories(c1)
bar.series[0].graphicalProperties.pattFill = None
wch.add_chart(bar, "D20")

# ── Top 10 sets bar — full name in col Q (17) so SUMIF never embeds a string ──
wch.cell(20, 15, "Set").font   = Font(bold=True)
wch.cell(20, 16, "USD").font   = Font(bold=True)
for k, (set_name, _) in enumerate(sorted_sets[:10], start=21):
    short = re.sub(r'^\d{4}(?:-\d{2})?\s*', '', set_name)[:32]
    wch.cell(k, 15, value=short)
    helper = wch.cell(k, 17, value=set_name)  # col Q — full name for SUMIF
    helper.font = Font(name="Calibri", size=9, color="AAAAAA")
    wch.cell(k, 16, value=f'=SUMIF({ac_range(C_SET)},Q{k},{ac_range(C_USD)})')
    wch.cell(k, 16).number_format = USD_FMT

bar2 = BarChart(); bar2.type = "bar"
bar2.title = "Top 10 Sets by Total Value (USD)"
bar2.style = 10; bar2.width = 18; bar2.height = 12
d2 = Reference(wch, min_col=16, min_row=20, max_row=30)
c2 = Reference(wch, min_col=15, min_row=21, max_row=30)
bar2.add_data(d2, titles_from_data=True); bar2.set_categories(c2)
bar2.series[0].graphicalProperties.pattFill = None
wch.add_chart(bar2, "P20")

wch.column_dimensions["A"].width = 28; wch.column_dimensions["B"].width = 14
wch.column_dimensions["O"].width = 28; wch.column_dimensions["P"].width = 14

# ════════════════════════════════════════════════════════════════════════════
# SHEET 6 — VALUE TIERS  (live counts + totals; card list links to All Cards)
# ════════════════════════════════════════════════════════════════════════════
wt = wb.create_sheet("Value Tiers")
wt.sheet_view.showGridLines = False
wt.column_dimensions["A"].width = 70
wt.column_dimensions["B"].width = 18
wt.column_dimensions["C"].width = 13

tier_colors = ["C00000","ED7D31","4472C4","70AD47","7030A0"]
current_row = 1
for ti, (label, lo, hi) in enumerate(tiers):
    hi_s = str(hi) if hi < 9_000_000 else "9999999"
    tc = tier_colors[ti]
    tier_list = sorted([rw for rw in priced if lo <= parse_price(rw["Fair Value"]) < hi],
                       key=lambda x: parse_price(x["Fair Value"]), reverse=True)

    wt.merge_cells(f"A{current_row}:C{current_row}")
    c = wt.cell(current_row, 1, value=label)
    c.font = Font(name="Calibri", bold=True, size=11, color=C_HEADER_FG)
    c.fill = fill(tc); c.alignment = Alignment(horizontal="center")
    wt.row_dimensions[current_row].height = 20; current_row += 1

    # Live count / total
    wt.cell(current_row, 1, value="Live Count:").font = Font(name="Calibri", bold=True, size=10)
    cnt_c = wt.cell(current_row, 2,
                    value=f'=COUNTIFS({ac_range(C_USD)},">={lo}",{ac_range(C_USD)},"<{hi_s}")')
    sum_c = wt.cell(current_row, 3,
                    value=f'=SUMIFS({ac_range(C_USD)},{ac_range(C_USD)},">={lo}",{ac_range(C_USD)},"<{hi_s}")')
    sum_c.number_format = USD_FMT
    cnt_c.alignment = Alignment(horizontal="center")
    sum_c.alignment = Alignment(horizontal="right")
    for ci in [1, 2, 3]:
        wt.cell(current_row, ci).fill = fill("FFFDE7")
        wt.cell(current_row, ci).border = border
    current_row += 1

    for ci, h in enumerate(["Card Name","Fair Value (USD)","Confidence"], start=1):
        c = wt.cell(current_row, ci, value=h)
        c.font = Font(name="Calibri", bold=True, size=10, color=C_HEADER_FG)
        c.fill = fill(C_HEADER_BG); c.border = border
    current_row += 1

    for rw in tier_list:
        ac_row = next((i + DATA_START for i, row in enumerate(rows)
                       if row["Card Name"] == rw["Card Name"]), None)
        bg = C_ALT if current_row % 2 == 0 else "FFFFFF"
        wt.cell(current_row, 1, value=rw["Card Name"])
        if ac_row:
            usd_ref = wt.cell(current_row, 2, value=f"='All Cards'!{col(C_USD)}{ac_row}")
            usd_ref.number_format = USD_FMT
            usd_ref.alignment = Alignment(horizontal="right")
        wt.cell(current_row, 3, value=rw.get("Confidence",""))
        wt.cell(current_row, 3).alignment = Alignment(horizontal="center")
        for ci in [1, 2, 3]:
            wt.cell(current_row, ci).fill = fill(bg)
            wt.cell(current_row, ci).border = border
            wt.cell(current_row, ci).font = Font(name="Calibri", size=10)
        current_row += 1
    current_row += 1  # spacer

# ════════════════════════════════════════════════════════════════════════════
# SHEET 7 — UNDER $5  (links to All Cards)
# ════════════════════════════════════════════════════════════════════════════
wu = wb.create_sheet("Under $5")
wu.sheet_view.showGridLines = False
wu.column_dimensions["A"].width = 70; wu.column_dimensions["B"].width = 18
wu.column_dimensions["C"].width = 13; wu.column_dimensions["D"].width = 8

hdr(wu, 1, f"CARDS UNDER $5  (snapshot list — values link to All Cards)", "C00000", cols=4, size=12)
wu.merge_cells("A2:D2"); wu.row_dimensions[2].height = 18
note_u = wu.cell(2, 1, value="Prices link to 'All Cards' sheet. If a price changes above $5, use the All Cards filter for a live view.")
note_u.font = Font(name="Calibri", italic=True, size=9, color="666666")

wu.cell(3, 1, value="Live count of cards under $5:").font = Font(name="Calibri", bold=True, size=10)
live_cnt = wu.cell(3, 2, value=f'=COUNTIFS({ac_range(C_USD)},">=0.01",{ac_range(C_USD)},"<5")')
live_cnt.alignment = Alignment(horizontal="center")
for ci in [1, 2]:
    wu.cell(3, ci).fill = fill(C_GOLD); wu.cell(3, ci).border = border

for ci, h in enumerate(["Card Name","Fair Value (USD)","Confidence","# Sales"], start=1):
    c = wu.cell(4, ci, value=h)
    c.font = Font(name="Calibri", bold=True, size=11, color=C_HEADER_FG)
    c.fill = fill(C_HEADER_BG); c.alignment = Alignment(horizontal="center"); c.border = border

for j, rw in enumerate(under5_rows, start=5):
    ac_row = next((i + DATA_START for i, row in enumerate(rows)
                   if row["Card Name"] == rw["Card Name"]), None)
    bg = C_ALT if j % 2 == 0 else "FFFFFF"
    wu.cell(j, 1, value=rw["Card Name"])
    if ac_row:
        usd_ref = wu.cell(j, 2, value=f"='All Cards'!{col(C_USD)}{ac_row}")
        usd_ref.number_format = USD_FMT
        usd_ref.alignment = Alignment(horizontal="right")
    wu.cell(j, 3, value=rw.get("Confidence","")).alignment = Alignment(horizontal="center")
    wu.cell(j, 4, value=rw.get("Num Sales","") or "—").alignment = Alignment(horizontal="center")
    for ci in [1, 2, 3, 4]:
        wu.cell(j, ci).fill = fill(bg); wu.cell(j, ci).border = border
        wu.cell(j, ci).font = Font(name="Calibri", size=10)

# ════════════════════════════════════════════════════════════════════════════
# SHEET 8 — MANUAL ENTRIES  (prices link to All Cards)
# ════════════════════════════════════════════════════════════════════════════
wm = wb.create_sheet("Manual Entries")
wm.sheet_view.showGridLines = False
wm.column_dimensions["A"].width = 70; wm.column_dimensions["B"].width = 20
wm.column_dimensions["C"].width = 40

hdr(wm, 1, f"MANUALLY PRICED CARDS  ({len(manual_rows)} cards) — prices link to All Cards", "7030A0", cols=3, size=12)
for ci, h in enumerate(["Card Name","Manual Price (USD)","Notes"], start=1):
    c = wm.cell(2, ci, value=h)
    c.font = Font(name="Calibri", bold=True, size=11, color=C_HEADER_FG)
    c.fill = fill(C_HEADER_BG); c.alignment = Alignment(horizontal="center"); c.border = border

for j, rw in enumerate(manual_rows, start=3):
    ac_row = next((i + DATA_START for i, row in enumerate(rows)
                   if row["Card Name"] == rw["Card Name"]), None)
    bg = C_ALT if j % 2 == 0 else "FFFFFF"
    wm.cell(j, 1, value=rw["Card Name"])
    if ac_row:
        usd_ref = wm.cell(j, 2, value=f"='All Cards'!{col(C_USD)}{ac_row}")
        usd_ref.number_format = USD_FMT
        usd_ref.alignment = Alignment(horizontal="right")
    wm.cell(j, 3, value="No eBay sold listings — comp price entered manually")
    for ci in [1, 2, 3]:
        wm.cell(j, ci).fill = fill(bg); wm.cell(j, ci).border = border
        wm.cell(j, ci).font = Font(name="Calibri", size=10)

man_tr = len(manual_rows) + 3
c = wm.cell(man_tr, 1, value="SUBTOTAL")
c.font = Font(name="Calibri", bold=True, size=11)
sub_usd = wm.cell(man_tr, 2, value=f'=SUMIF({ac_range(C_CONF)},"manual",{ac_range(C_USD)})')
sub_usd.number_format = USD_FMT
for ci in [1, 2]:
    c = wm.cell(man_tr, ci); c.fill = fill(C_MANUAL); c.border = border
    c.font = Font(name="Calibri", bold=True, size=11)
sub_usd.alignment = Alignment(horizontal="right")

# ── Save ────────────────────────────────────────────────────────────────────
out = "batch_price_report.xlsx"
wb.save(out)
print(f"Saved: {out}")
print(f"Sheets: All Cards | Summary | By Player | By Set | Charts | Value Tiers | Under $5 | Manual Entries")
print(f"\nEditing: go to All Cards sheet, edit any yellow cell in column C (Fair Value USD)")
print(f"All summaries, player/set totals, and charts update automatically.")
